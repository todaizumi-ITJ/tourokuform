"""
新規案件 通知判定 Webアプリ
Flask + BeautifulSoup4 + openpyxl
v8ロジック:
  - IPNなし → 一律通知しない
  - IPNあり → STEP1氏名照合 → マッチなし→通知する
  - マッチあり → STEP3: 受任済み×同一CP → 通知しない
  - マッチあり → STEP4: 包括和解済み（hn_id × cp_num）→ 通知しない
  - それ以外 → 通知する
"""
import json, unicodedata, re, os, io, time
from datetime import datetime
from collections import defaultdict
from flask import Flask, request, jsonify, send_file, render_template
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

# ========================
# DBロード（起動時に1回）
# ========================
def load_db():
    t0 = time.time()

    # 既存案件DB
    with open(os.path.join(DATA_DIR, 'db_human.json'), encoding='utf-8') as f:
        human = json.load(f)

    # IPN → {cp_num, cw_num, rh} 対応表
    with open(os.path.join(DATA_DIR, 'db_ipn.json'), encoding='utf-8') as f:
        ipn_map = json.load(f)

    # 包括和解済み（hn_id × cp_num のセット）
    with open(os.path.join(DATA_DIR, 'db_settled.json'), encoding='utf-8') as f:
        settled_data = json.load(f)

    # IP+タイムスタンプ → レコードリスト（IPN検索用）
    with open(os.path.join(DATA_DIR, 'db_ip_lookup.json'), encoding='utf-8') as f:
        ip_lookup = json.load(f)

    # settled_set: {(hn_id, cp_num), ...}
    settled_set = {(r['hn_id'], r['cp_num']) for r in settled_data}

    # 氏名インデックス（正規化後 → [rec, ...]）
    name_idx = defaultdict(list)
    for rec in human:
        n = norm_name(rec.get('contract_name') or rec.get('user_name', ''))
        if n:
            name_idx[n].append(rec)

    elapsed = time.time() - t0
    print(f"DB loaded: human={len(human):,} ipn={len(ipn_map):,} settled={len(settled_set)} ip_lookup={len(ip_lookup):,} ({elapsed:.2f}s)")
    return human, ipn_map, settled_set, name_idx, ip_lookup

# ========================
# 正規化
# ========================
VARIANT_MAP = str.maketrans({
    '斎':'斉','齋':'斉','齊':'斉','邉':'辺','邊':'辺',
    '髙':'高','﨑':'崎','嶋':'島','嶌':'島','濱':'浜','濵':'浜',
    '瀧':'滝','廣':'広','澤':'沢','眞':'真','國':'国','德':'徳',
    '桒':'桑','栁':'柳','龍':'竜','關':'関',
})

def norm_name(n):
    if not n: return ''
    n = unicodedata.normalize('NFKC', n)
    n = re.sub(r'[\s\u3000\u00a0]+', '', n)
    return n.translate(VARIANT_MAP)

def norm_addr(a):
    if not a: return ''
    a = unicodedata.normalize('NFKC', a)
    a = a.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
    a = re.sub(r'[\s\u3000]+', '', a)
    a = re.sub(r'([0-9]+)丁目([0-9]+)番地?([0-9]+)号?', r'\1-\2-\3', a)
    a = re.sub(r'丁目', '-', a)
    a = re.sub(r'番地?', '-', a)
    a = re.sub(r'号', '', a)
    a = re.sub(r'[-－ー]+', '-', a)
    return a[:40]

# ========================
# HTMLパース
# ========================
def _c(cells, i):
    return cells[i].strip() if len(cells) > i else ''

def parse_html(content: bytes, filename: str) -> list:
    soup = BeautifulSoup(content, 'html.parser')
    records = []
    rows = soup.find_all('tr')
    for row in rows[2:]:
        cells = [td.get_text(strip=True) for td in row.find_all(['td', 'th'])]
        if len(cells) < 8: continue
        name = cells[2].strip()
        addr = cells[7].strip() if len(cells) > 7 else ''
        if not name and not addr: continue
        records.append({
            'source_file':          filename,
            'registration_type':    _c(cells, 1),
            'contract_name':        _c(cells, 2),
            'contract_name_kana':   _c(cells, 3),
            'user_name':            _c(cells, 4),
            'user_name_kana':       _c(cells, 5),
            'postal_code':          _c(cells, 6),
            'address':              _c(cells, 7),
            'phone':                _c(cells, 8),
            'email':                _c(cells, 9),
            'ipn_id':               _c(cells, 10),
            'ip_address':           _c(cells, 11),
            'port':                 _c(cells, 12),
            'timestamp':            _c(cells, 13),
            'provider':             _c(cells, 14),
            'lawyer_id':            _c(cells, 15),
            'memo':                 _c(cells, 16),
            'rights_holder_new':    _c(cells, 17),
            'google_drive_url':     _c(cells, 18),
            'pdf_filename':         _c(cells, 19),
        })
    return records

# ========================
# 通知判定ロジック（v8）
# ========================
def judge(new: dict, name_idx: dict, ipn_map: dict, settled_set: set) -> dict:
    nn  = norm_name(new['contract_name'])
    na  = norm_addr(new['address'])
    ipn = new.get('ipn_id', '').strip()

    # IPNなし → 要確認（一律抑制せず人が判断）
    if not ipn:
        rh_new = new.get('rights_holder', '') or new.get('rights_holder_new', '')
        reason = f'IPNなし・要確認（権利者: {rh_new}）' if rh_new else 'IPNなし・要確認'
        return _result(new, '⚠️ 要確認', reason, None, rh_new, None, '', [])

    # IPN → cp_num / rh を解決
    cp_info = ipn_map.get(ipn)
    if not cp_info:
        return _result(new, '❌ 通知しない', f'IPN未解決（{ipn}）・通知しない',
                       None, '', None, '', [])

    cp_num = cp_info['cp_num']
    rh     = cp_info.get('rh', '')
    cw_num = cp_info.get('cw_num', '')

    # STEP1: 氏名照合
    candidates = name_idx.get(nn, [])

    if not candidates:
        # マッチなし → 完全新規 → 通知する
        return _result(new, '✅ 通知する', f'新規（既存DBにマッチなし）/ {rh}',
                       cp_num, rh, cw_num, '', [])

    # 最良マッチ（住所スコア）
    best_rec, best_score = _best_match(candidates, na)
    addr_label = {2: '住所一致（前20文字）', 1: '住所部分一致（前12文字）', 0: '住所不一致'}[best_score]

    # STEP3: 受任済み × 同一CP（cp_numで数値照合）
    mandated = [r for r in candidates if r.get('mandate_status') == '受任済']
    same_cp_mandated = [r for r in mandated if cp_num in (r.get('cp_nums') or [])]
    if same_cp_mandated:
        reason = f'条件①: 受任済み×同一CP（{rh} / cp_num={cp_num}）/ {len(same_cp_mandated)}件'
        return _result(new, '❌ 通知しない', reason,
                       cp_num, rh, cw_num, addr_label,
                       _candidates_detail(candidates, na))

    # STEP4: 包括和解済み（hn_id × cp_num）
    # STEP1でマッチした既存HNのうち、同一CPで包括和解済みのものがあるか
    settled_matched = [r for r in candidates if (r.get('hn_id'), cp_num) in settled_set]
    if settled_matched:
        hn_list = ', '.join(r['hn_id'] for r in settled_matched)
        reason = f'条件②: 包括和解済み（{rh} / cp_num={cp_num} / HN: {hn_list}）'
        return _result(new, '❌ 通知しない', reason,
                       cp_num, rh, cw_num, addr_label,
                       _candidates_detail(candidates, na))

    # 条件①②非該当 → 通知する
    return _result(new, '✅ 通知する',
                   f'既存あり・条件①②非該当（{rh} / cp_num={cp_num}）',
                   cp_num, rh, cw_num, addr_label,
                   _candidates_detail(candidates, na))


def _best_match(candidates, na):
    best_rec = candidates[0]; best_score = 0
    for rec in candidates:
        ea = norm_addr(rec.get('address', ''))
        s = 2 if (na and ea and na[:20] == ea[:20]) else \
            1 if (na and ea and na[:12] == ea[:12]) else 0
        if s > best_score:
            best_score = s; best_rec = rec
    return best_rec, best_score

def _addr_score_label(na, rec):
    ea = norm_addr(rec.get('address', ''))
    if na and ea and na[:20] == ea[:20]: return '住所一致（前20文字）'
    if na and ea and na[:12] == ea[:12]: return '住所部分一致（前12文字）'
    return '住所不一致'

def _candidates_detail(candidates, na):
    details = []
    for rec in candidates:
        details.append({
            'hn_id':             rec.get('hn_id', ''),
            'contract_name':     rec.get('contract_name', ''),
            'address':           rec.get('address', ''),
            'mandate_status':    rec.get('mandate_status', ''),
            'settlement_status': rec.get('settlement_status', ''),
            'rights_holder':     rec.get('rights_holder', ''),
            'cp_nums':           rec.get('cp_nums', []),
            'kmp_progress':      rec.get('kmp_progress', ''),
            'lawyer_name':       rec.get('lawyer_name', ''),
            'lawyer_office':     rec.get('lawyer_office', ''),
            'addr_match':        _addr_score_label(na, rec),
        })
    return details

def _result(new, judgment, reason, cp_num, rh, cw_num, addr_label, candidates_detail):
    best_rec = None
    if candidates_detail:
        # addr_matchスコアが最高のものを代表として使う
        score_map = {'住所一致（前20文字）': 2, '住所部分一致（前12文字）': 1, '住所不一致': 0}
        best_cand = max(candidates_detail, key=lambda c: score_map.get(c.get('addr_match', ''), 0))
    else:
        best_cand = {}

    return {
        '通知判定':             judgment,
        '判定理由':             reason,
        'source_file':         new.get('source_file', ''),
        # 新規側フィールド
        'registration_type':   new.get('registration_type', ''),
        'contract_name':       new.get('contract_name', ''),
        'contract_name_kana':  new.get('contract_name_kana', ''),
        'user_name':           new.get('user_name', ''),
        'user_name_kana':      new.get('user_name_kana', ''),
        'postal_code':         new.get('postal_code', ''),
        'address':             new.get('address', ''),
        'phone':               new.get('phone', ''),
        'email':               new.get('email', ''),
        'ipn_id':              new.get('ipn_id', ''),
        'ip_address':          new.get('ip_address', ''),
        'port':                new.get('port', ''),
        'timestamp':           new.get('timestamp', ''),
        'provider':            new.get('provider', ''),
        'lawyer_id':           new.get('lawyer_id', ''),
        'memo':                new.get('memo', ''),
        'rights_holder_new':   new.get('rights_holder_new', ''),
        'google_drive_url':    new.get('google_drive_url', ''),
        'pdf_filename':        new.get('pdf_filename', ''),
        # CP/CW解決結果
        'cp_num':              cp_num,
        'cw_num':              cw_num,
        'rights_holder_ipn':   rh,
        # 既存DB照合結果（最良候補）
        '既存_HN_ID':          best_cand.get('hn_id', ''),
        '既存_氏名':           best_cand.get('contract_name', ''),
        '既存_住所':           best_cand.get('address', ''),
        '既存_受任状況':       best_cand.get('mandate_status', ''),
        '既存_和解状況':       best_cand.get('settlement_status', ''),
        '既存_KMP進捗':        best_cand.get('kmp_progress', ''),
        '既存_権利者':         best_cand.get('rights_holder', ''),
        '既存_法律事務所':     best_cand.get('lawyer_office', ''),
        '住所一致度':          addr_label,
        'candidates_detail':   candidates_detail,
    }

# ========================
# Excel生成
# ========================
def make_excel(results: list) -> bytes:
    wb = openpyxl.Workbook()

    C_H = '1F3864'; C_HF = 'FFFFFF'
    C_OK_B = 'C6EFCE'; C_OK_F = '276221'
    C_NG_B = 'FFCCCC'; C_NG_F = '9C0006'
    C_NG2_B = 'FFE0CC'; C_NG2_F = '7B3000'
    C_WN_B = 'FFEB9C'; C_WN_F = '9C6500'   # ⚠️ 要確認 (黄)
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'))

    def hdr(cell, text):
        cell.value = text
        cell.font = Font(name='Arial', bold=True, color=C_HF, size=10)
        cell.fill = PatternFill('solid', fgColor=C_H)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin

    def dat(cell, value='', bg=None):
        cell.value = value
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if bg: cell.fill = PatternFill('solid', fgColor=bg)
        cell.border = thin

    def nc(cell, val, reason=''):
        is_ok = str(val).startswith('✅')
        is_wn = str(val).startswith('⚠️')
        is_s2 = '条件②' in str(reason)
        cell.value = val
        if is_ok:
            cell.font = Font(name='Arial', bold=True, color=C_OK_F, size=10)
            cell.fill = PatternFill('solid', fgColor=C_OK_B)
        elif is_wn:
            cell.font = Font(name='Arial', bold=True, color=C_WN_F, size=10)
            cell.fill = PatternFill('solid', fgColor=C_WN_B)
        elif is_s2:
            cell.font = Font(name='Arial', bold=True, color=C_NG2_F, size=10)
            cell.fill = PatternFill('solid', fgColor=C_NG2_B)
        else:
            cell.font = Font(name='Arial', bold=True, color=C_NG_F, size=10)
            cell.fill = PatternFill('solid', fgColor=C_NG_B)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin

    COLS = [
        ('通知判定',14),('判定理由',48),('ソースファイル',12),
        # 新規側
        ('登録種別',10),('氏名',14),('カナ',18),('ユーザー名',14),('ユーザーカナ',18),
        ('郵便番号',10),('住所',30),('電話',18),('メール',26),
        ('IPN',12),('IPアドレス',16),('ポート',8),('タイムスタンプ',18),
        ('プロバイダ',20),('弁護士ID',10),('メモ',20),
        ('新規_権利者',22),('Google_Drive_URL',22),('PDFファイル名',20),
        # CP/CW解決
        ('CP番号',8),('CW番号',8),('権利者(IPN解決)',22),
        # 既存DB照合
        ('既存_HN_ID',12),('既存_氏名',14),('既存_住所',28),
        ('既存_受任状況',12),('既存_和解状況',12),('既存_KMP進捗',12),('既存_権利者',22),
        ('既存_法律事務所',28),
        ('住所一致度',18),('同姓同名件数',10),
    ]

    def build_sheet(ws, rows, title):
        ws.title = title
        for i,(name,w) in enumerate(COLS,1):
            ws.column_dimensions[get_column_letter(i)].width = w
            hdr(ws.cell(1,i), name)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'
        for ri,r in enumerate(rows,2):
            reason = r.get('判定理由','')
            is_ok = str(r.get('通知判定','')).startswith('✅')
            is_wn = str(r.get('通知判定','')).startswith('⚠️')
            is_s2 = '条件②' in reason
            if is_ok:
                bg = 'EAF7EA' if ri%2==0 else 'F5FFF5'
            elif is_wn:
                bg = 'FFFACD' if ri%2==0 else 'FFFFF0'
            elif is_s2:
                bg = 'FFE8D0' if ri%2==0 else 'FFF4EC'
            else:
                bg = 'FFF5F5' if ri%2==0 else 'FFFFFF'
            cand_count = len(r.get('candidates_detail', []))
            vals = [
                r.get('通知判定',''), reason, r.get('source_file',''),
                r.get('registration_type',''), r.get('contract_name',''), r.get('contract_name_kana',''),
                r.get('user_name',''), r.get('user_name_kana',''),
                r.get('postal_code',''), r.get('address',''), r.get('phone',''), r.get('email',''),
                r.get('ipn_id',''), r.get('ip_address',''), r.get('port',''), r.get('timestamp',''),
                r.get('provider',''), r.get('lawyer_id',''), r.get('memo',''),
                r.get('rights_holder_new',''), r.get('google_drive_url',''), r.get('pdf_filename',''),
                str(r['cp_num']) if r.get('cp_num') is not None else '',
                str(r['cw_num']) if r.get('cw_num') not in (None, '') else '',
                r.get('rights_holder_ipn',''),
                r.get('既存_HN_ID',''), r.get('既存_氏名',''), r.get('既存_住所',''),
                r.get('既存_受任状況',''), r.get('既存_和解状況',''), r.get('既存_KMP進捗',''), r.get('既存_権利者',''),
                r.get('既存_法律事務所',''),
                r.get('住所一致度',''), cand_count if cand_count else '',
            ]
            for ci,val in enumerate(vals,1):
                cell = ws.cell(ri,ci,val)
                if ci == 1: nc(cell, val, reason)
                else: dat(cell, val, bg=bg)
            ws.row_dimensions[ri].height = 18

    def build_ng_detail_sheet(ws, ng_rows):
        ws.title = '❌ 通知しない（詳細）'
        DCOLS = [
            ('新規_氏名',14),('新規_住所',28),('新規_IPN',10),('新規_CP権利者',22),
            ('判定理由',48),
            ('既存_HN_ID',12),('既存_氏名',14),('既存_住所',28),
            ('既存_受任状況',12),('既存_和解状況',12),('既存_KMP進捗',12),('既存_権利者',22),
            ('法律事務所',28),('担当弁護士',16),
            ('住所一致度',18),
        ]
        for i,(name,w) in enumerate(DCOLS,1):
            ws.column_dimensions[get_column_letter(i)].width = w
            hdr(ws.cell(1,i), name)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

        ri = 2
        for r in ng_rows:
            candidates = r.get('candidates_detail', [])
            reason = r.get('判定理由', '')
            new_name = r.get('contract_name','')
            new_addr = r.get('address','')
            new_ipn  = r.get('ipn_id','')
            new_rh   = r.get('rights_holder_ipn','')

            if not candidates:
                candidates = [{
                    'hn_id': r.get('既存_HN_ID',''),
                    'contract_name': r.get('既存_氏名',''),
                    'address': r.get('既存_住所',''),
                    'mandate_status': r.get('既存_受任状況',''),
                    'settlement_status': r.get('既存_和解状況',''),
                    'rights_holder': r.get('既存_権利者',''),
                    'kmp_progress': r.get('既存_KMP進捗',''),
                    'lawyer_office': r.get('既存_法律事務所',''),
                    'lawyer_name': '',
                    'addr_match': r.get('住所一致度',''),
                }]

            for cand in candidates:
                is_mandated = cand.get('mandate_status') == '受任済'
                bg_row = 'FFF0F0' if is_mandated else 'FFFFFF'
                vals = [
                    new_name, new_addr, new_ipn, new_rh, reason,
                    cand.get('hn_id',''), cand.get('contract_name',''), cand.get('address',''),
                    cand.get('mandate_status',''), cand.get('settlement_status',''),
                    cand.get('kmp_progress',''), cand.get('rights_holder',''),
                    cand.get('lawyer_office',''), cand.get('lawyer_name',''),
                    cand.get('addr_match',''),
                ]
                for ci,val in enumerate(vals,1):
                    cell = ws.cell(ri,ci,val)
                    dat(cell, val, bg=bg_row)
                    if ci == 9 and is_mandated:
                        cell.font = Font(name='Arial', bold=True, color='9C0006', size=10)
                ws.row_dimensions[ri].height = 18
                ri += 1

    def sort_key(r):
        j = str(r.get('通知判定',''))
        if j.startswith('✅'): return (0, r.get('contract_name',''))
        if j.startswith('⚠️'): return (1, r.get('contract_name',''))
        return (2, r.get('contract_name',''))

    sorted_all = sorted(results, key=sort_key)
    ok_rows = [r for r in sorted_all if str(r.get('通知判定','')).startswith('✅')]
    wn_rows = [r for r in sorted_all if str(r.get('通知判定','')).startswith('⚠️')]
    ng_rows = [r for r in sorted_all if str(r.get('通知判定','')).startswith('❌')]

    ws0 = wb.active
    build_sheet(ws0, sorted_all, '📋 全件判定結果')
    build_sheet(wb.create_sheet(), ok_rows, '✅ 通知する')
    build_sheet(wb.create_sheet(), wn_rows, '⚠️ 要確認（IPNなし）')
    build_sheet(wb.create_sheet(), ng_rows, '❌ 通知しない')
    build_ng_detail_sheet(wb.create_sheet(), ng_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ========================
# グローバルDB（起動時ロード）
# ========================
HUMAN, IPN_MAP, SETTLED_SET, NAME_IDX, IP_LOOKUP = load_db()

# ========================
# ルーティング
# ========================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/lookup')
def lookup():
    return render_template('lookup.html')

@app.route('/api/judge', methods=['POST'])
def api_judge():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'ファイルが見つかりません'}), 400

    all_records = []
    for f in files:
        try:
            recs = parse_html(f.read(), f.filename)
            all_records.extend(recs)
        except Exception as e:
            return jsonify({'error': f'{f.filename}: {str(e)}'}), 400

    results = [judge(rec, NAME_IDX, IPN_MAP, SETTLED_SET) for rec in all_records]

    from collections import Counter
    jc = Counter(r['通知判定'] for r in results)
    rc = Counter(r['判定理由'] for r in results)

    return jsonify({
        'total':    len(results),
        'notify':   jc.get('✅ 通知する', 0),
        'suppress': jc.get('❌ 通知しない', 0),
        'reasons':  dict(rc),
        'results':  results,
    })

@app.route('/api/export', methods=['POST'])
def api_export():
    data = request.get_json()
    results = data.get('results', [])
    if not results:
        return jsonify({'error': 'データなし'}), 400

    xlsx_bytes = make_excel(results)
    fname = f'通知判定結果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )

@app.route('/api/thn/batch', methods=['POST'])
def api_thn_batch():
    """
    THN JSON / Staging JSON を直接受け取って通知判定する。
    入力: [{contract_name, address, ipn_id, rights_holder, ...}, ...]
    _meta フィールドは無視される（staging互換）。
    """
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'JSONが空です'}), 400
    if not isinstance(data, list):
        data = [data]

    results = []
    for entry in data:
        # _meta は判定に使わず、source情報だけ引き継ぐ
        meta   = entry.get('_meta', {})
        source = entry.get('source_type', 'provider')

        # staging / THN 両フォーマット対応の正規化
        rec = {
            'contract_name':      (entry.get('contract_name') or entry.get('user_name') or '').strip(),
            'contract_name_kana': (entry.get('contract_name_kana') or entry.get('user_name_kana') or '').strip(),
            'address':            (entry.get('address') or '').strip(),
            'postal_code':        (entry.get('postal_code') or '').strip(),
            'phone':              (entry.get('phone') or '').strip(),
            'email':              (entry.get('email') or '').strip(),
            'ipn_id':             (entry.get('ipn_id') or '').strip(),
            'ip_address':         (entry.get('ip_address') or '').strip(),
            'port':               (entry.get('port') or '').strip(),
            'timestamp':          (entry.get('timestamp') or '').strip(),
            'provider':           (entry.get('provider') or '').strip(),
            'rights_holder':      (entry.get('rights_holder') or '').strip(),
            'rights_holder_new':  (entry.get('rights_holder') or '').strip(),
            'lawyer_id':          (entry.get('lawyer_id') or '').strip(),
            'memo':               (entry.get('memo') or '').strip(),
            'google_drive_url':   (entry.get('google_drive_url') or '').strip(),
            'pdf_filename':       (entry.get('pdf_filename') or '').strip(),
            'registration_type':  (entry.get('registration_type') or '').strip(),
            'source_file':        entry.get('thn_id') or f"row_{meta.get('row_number','')}",
            # _meta情報を引き継ぎ
            'thn_id':             entry.get('thn_id') or f"THN-{meta.get('batch_id','')}-{meta.get('row_number','')}",
            'thn_status':         entry.get('thn_status') or meta.get('status', 'pending'),
            'batch_id':           meta.get('batch_id', ''),
            'row_number':         meta.get('row_number', ''),
            'error_message':      meta.get('error_message', ''),
        }
        results.append(judge(rec, NAME_IDX, IPN_MAP, SETTLED_SET))

    from collections import Counter
    jc = Counter(r['通知判定'] for r in results)

    return jsonify({
        'total':    len(results),
        'notify':   jc.get('✅ 通知する', 0),
        'warn':     jc.get('⚠️ 要確認', 0),
        'suppress': jc.get('❌ 通知しない', 0),
        'results':  results,
    })


@app.route('/api/lookup', methods=['POST'])
def api_lookup():
    """IP+タイムスタンプ CSVからIPN検索"""
    data = request.get_json()
    rows = data.get('rows', [])  # [{ip, timestamp}, ...]
    results = []
    not_found = []
    for row in rows:
        ip = (row.get('ip') or '').strip()
        ts = (row.get('timestamp') or '').strip()
        if not ip or not ts:
            continue
        key = f'{ip}|{ts}'
        hits = IP_LOOKUP.get(key, [])
        if hits:
            results.append({'ip': ip, 'timestamp': ts, 'records': hits})
        else:
            not_found.append({'ip': ip, 'timestamp': ts})
    return jsonify({
        'found': len(results),
        'not_found': len(not_found),
        'results': results,
        'not_found_list': not_found,
    })

@app.route('/api/reload-db', methods=['POST'])
def api_reload_db():
    global HUMAN, IPN_MAP, SETTLED_SET, NAME_IDX, IP_LOOKUP
    try:
        HUMAN, IPN_MAP, SETTLED_SET, NAME_IDX, IP_LOOKUP = load_db()
        return jsonify({
            'message': f'DB再読み込み完了: human={len(HUMAN):,}件, ipn={len(IPN_MAP):,}件, settled={len(SETTLED_SET)}件, ip_lookup={len(IP_LOOKUP):,}件'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-status', methods=['GET'])
def api_db_status():
    return jsonify({
        'human_count':   len(HUMAN),
        'ipn_count':     len(IPN_MAP),
        'settled_count': len(SETTLED_SET),
        'settled_list':  [{'hn_id': h, 'cp_num': c} for h, c in sorted(SETTLED_SET)],
    })

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
